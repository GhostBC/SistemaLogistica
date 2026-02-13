"""Exportação de relatórios para Excel."""
from datetime import date, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Formato moeda para células
FMT_MOEDA = '"R$ "#,##0.00'
FMT_PERCENTUAL = '0.00"%"'
FMT_DATA = 'DD/MM/YYYY'
FILL_CABECALHO = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def _texto_embalagem(p):
    """Monta texto da coluna Embalagem a partir do pedido (dict)."""
    if p.get('embalagens') and len(p['embalagens']) > 0:
        partes = []
        for pe in p['embalagens']:
            nome = (pe.get('embalagem') or {}).get('nome', 'Desconhecida') if isinstance(pe.get('embalagem'), dict) else 'Desconhecida'
            qtd = pe.get('quantidade', 1)
            partes.append('{} (x{})'.format(nome, qtd))
        return ', '.join(partes)
    if p.get('embalagem') and isinstance(p['embalagem'], dict):
        nome = p['embalagem'].get('nome', '-')
        qtd = p.get('quantidade_embalagem', 1)
        return '{} (x{})'.format(nome, qtd)
    return '-'


def exportar_finalizados(pedidos_list):
    """
    Gera arquivo Excel com a lista de pedidos finalizados (mesmas colunas da aba Finalizados).
    pedidos_list: list of dicts (numero_pedido, loja_nome, marketplace, frete_cliente, peso, transportadora, custo_mandae, tracking_code, data_finalizacao, embalagens/embalagem)
    Retorna: BytesIO com o arquivo .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pedidos Finalizados'

    ws['A1'] = 'Exportação - Pedidos Finalizados'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:J1')
    row = 3

    headers = ['Nº Pedido', 'Loja', 'Marketplace', 'Frete (R$)', 'Peso (kg)', 'Transportadora', 'Frete real (R$)', 'Embalagem', 'Rastreio', 'Data finalização']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = FILL_CABECALHO
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    row += 1

    for p in pedidos_list:
        data_fim = p.get('data_finalizacao')
        if data_fim:
            try:
                if isinstance(data_fim, str):
                    dt = datetime.fromisoformat(data_fim.replace('Z', '+00:00'))
                    data_fim = dt.strftime('%d/%m/%Y %H:%M') if dt else '-'
                else:
                    data_fim = data_fim.strftime('%d/%m/%Y %H:%M')
            except Exception:
                data_fim = str(data_fim)
        else:
            data_fim = '-'

        ws.cell(row=row, column=1, value=p.get('numero_pedido') or '-').border = thin_border
        ws.cell(row=row, column=2, value=p.get('loja_nome') or '-').border = thin_border
        ws.cell(row=row, column=3, value=p.get('marketplace') or '-').border = thin_border
        cell_frete = ws.cell(row=row, column=4)
        cell_frete.value = round(float(p.get('frete_cliente') or 0), 2)
        cell_frete.number_format = FMT_MOEDA
        cell_frete.border = thin_border
        cell_peso = ws.cell(row=row, column=5)
        val_peso = p.get('peso')
        cell_peso.value = round(float(val_peso), 2) if val_peso is not None else '-'
        cell_peso.border = thin_border
        ws.cell(row=row, column=6, value=p.get('transportadora') or '-').border = thin_border
        cell_real = ws.cell(row=row, column=7)
        cell_real.value = round(float(p.get('custo_mandae') or 0), 2)
        cell_real.number_format = FMT_MOEDA
        cell_real.border = thin_border
        ws.cell(row=row, column=8, value=_texto_embalagem(p)).border = thin_border
        ws.cell(row=row, column=9, value=p.get('tracking_code') or '-').border = thin_border
        ws.cell(row=row, column=10, value=data_fim).border = thin_border
        row += 1

    for col, width in [(1, 14), (2, 12), (3, 14), (4, 12), (5, 10), (6, 14), (7, 14), (8, 28), (9, 18), (10, 18)]:
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def exportar_relatorio_periodo(inicio, fim, dados_consolidados):
    """
    Gera arquivo Excel do relatório por período.
    inicio, fim: date
    dados_consolidados: dict com totais e por_dia
    Retorna: BytesIO com o arquivo .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Relatório {inicio} a {fim}"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalho
    ws['A1'] = f'Relatório de Logística - {inicio} a {fim}'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    row = 3
    ws[f'A{row}'] = 'Resumo do Período'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    for key, label in [
        ('total_pedidos', 'Total de pedidos'),
        ('custo_total', 'Custo total (R$)'),
        ('frete_total', 'Pago cliente (R$)'),
        ('frete_real_total', 'Frete real total (R$)'),
        ('ganho_perda_liquido', 'Ganho/Perda líquido (R$)'),
    ]:
        ws[f'A{row}'] = label
        valor = dados_consolidados.get(key, 0)
        cell_b = ws[f'B{row}']
        if key in ('custo_total', 'frete_total', 'frete_real_total', 'ganho_perda_liquido'):
            cell_b.value = round(float(valor or 0), 2)
            cell_b.number_format = FMT_MOEDA
        else:
            cell_b.value = valor
        row += 1

    # Embalagens utilizadas
    if dados_consolidados.get('embalagens_utilizadas'):
        row += 1
        ws[f'A{row}'] = 'Embalagens Utilizadas'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = 'Embalagem'
        ws[f'B{row}'] = 'Quantidade'
        ws[f'C{row}'] = 'Custo Unitário (R$)'
        ws[f'D{row}'] = 'Valor Total (R$)'
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = FILL_CABECALHO
        row += 1
        
        for emb in dados_consolidados['embalagens_utilizadas']:
            ws[f'A{row}'] = emb.get('nome', '-')
            ws[f'B{row}'] = emb.get('quantidade', 0)
            c_c, d_c = ws[f'C{row}'], ws[f'D{row}']
            c_c.value = round(float(emb.get('custo_unitario') or 0), 2)
            c_c.number_format = FMT_MOEDA
            vt_emb = emb.get('valor_total') or (emb.get('quantidade', 0) * (emb.get('custo_unitario') or 0))
            d_c.value = round(float(vt_emb), 2)
            d_c.number_format = FMT_MOEDA
            for cell in [ws[f'A{row}'], ws[f'B{row}'], c_c, d_c]:
                cell.border = thin_border
            row += 1

    # Detalhamento por dia
    if dados_consolidados.get('por_dia'):
        row += 2
        ws[f'A{row}'] = 'Detalhamento por Dia'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = 'Data'
        ws[f'B{row}'] = 'Pedidos'
        ws[f'C{row}'] = 'Custo Total (R$)'
        ws[f'D{row}'] = 'Pago Cliente (R$)'
        ws[f'E{row}'] = 'Frete Real (R$)'
        ws[f'F{row}'] = 'Ganho/Perda (R$)'
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}'], ws[f'E{row}'], ws[f'F{row}']]:
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = FILL_CABECALHO
        row += 1
        
        for dia in dados_consolidados['por_dia']:
            ws[f'A{row}'] = dia.get('data', '-')
            ws[f'B{row}'] = dia.get('total_pedidos', 0)
            for col, key in enumerate(['custo_total', 'frete_total', 'frete_real_total', 'ganho_perda_liquido'], 3):
                cell = ws.cell(row=row, column=col)
                cell.value = round(float(dia.get(key) or 0), 2)
                cell.number_format = FMT_MOEDA
                cell.border = thin_border
            for col in [1, 2]:
                ws.cell(row=row, column=col).border = thin_border
            row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.freeze_panes = 'A4'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def exportar_relatorio_diario(data_relatorio, dados_consolidados, custos_lista):
    """
    Gera arquivo Excel do relatório diário.
    data_relatorio: date
    dados_consolidados: dict com total_pedidos, custo_total, ganho_total, etc
    custos_lista: list of CustoFrete ou dicts
    Retorna: BytesIO com o arquivo .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Relatório {data_relatorio}"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalho
    ws['A1'] = f'Relatório de Logística - {data_relatorio}'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')

    row = 3
    ws[f'A{row}'] = 'Resumo do dia'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    moeda_keys = ('custo_total', 'frete_total', 'ganho_total', 'perda_total')
    for key, label in [
        ('total_pedidos', 'Total de pedidos'),
        ('custo_total', 'Custo total (R$)'),
        ('frete_total', 'Pago cliente (R$)'),
        ('ganho_total', 'Ganho total (R$)'),
        ('perda_total', 'Perda total (R$)'),
    ]:
        ws[f'A{row}'] = label
        valor = dados_consolidados.get(key, 0)
        cell_b = ws[f'B{row}']
        if key in moeda_keys:
            cell_b.value = round(float(valor or 0), 2)
            cell_b.number_format = FMT_MOEDA
        else:
            cell_b.value = valor
        cell_b.border = thin_border
        row += 1

    row += 1
    embalagens_utilizadas = dados_consolidados.get('embalagens_utilizadas') or []
    if embalagens_utilizadas:
        ws[f'A{row}'] = 'Embalagens utilizadas no dia'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = 'Embalagem'
        ws[f'B{row}'] = 'Quantidade'
        ws[f'C{row}'] = 'Custo Unitário (R$)'
        ws[f'D{row}'] = 'Valor Total (R$)'
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = FILL_CABECALHO
        row += 1
        for eu in embalagens_utilizadas:
            vt = eu.get('valor_total') or (eu.get('quantidade', 0) * (eu.get('custo_unitario') or 0))
            ws[f'A{row}'] = eu.get('nome', '-')
            ws[f'B{row}'] = eu.get('quantidade', 0)
            ws[f'C{row}'] = round(float(eu.get('custo_unitario') or 0), 2)
            ws[f'C{row}'].number_format = FMT_MOEDA
            ws[f'D{row}'] = round(float(vt), 2)
            ws[f'D{row}'].number_format = FMT_MOEDA
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = thin_border
            row += 1
        row += 1

    widths = [14, 16, 14, 16, 14, 16, 12]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A4'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def exportar_relatorio_por_canal(inicio, fim, resultado):
    """
    Gera arquivo Excel do relatório por canal.
    inicio, fim: date (ou strings ISO)
    resultado: dict com inicio, fim, canais (list de canal, total_pedidos, custo_total, frete_total, frete_real_total, ganho_perda_liquido, caixas)
    Retorna: BytesIO com o arquivo .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Por Canal'

    # Título
    ws['A1'] = f'Relatório por Canal - {inicio} a {fim}'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    row = 3

    # Tabela Pedidos: cabeçalho em linha destacada
    headers_pedidos = ['Canal', 'Pedidos', 'Custo total (R$)', 'Pago cliente (R$)', 'Frete real (R$)', 'Ganho/Perda (R$)']
    for col, h in enumerate(headers_pedidos, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = FILL_CABECALHO
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    row += 1

    canais = resultado.get('canais') or []
    for c in canais:
        ws.cell(row=row, column=1, value=c.get('canal') or '-').border = thin_border
        ws.cell(row=row, column=2, value=c.get('total_pedidos') or 0).border = thin_border
        for col, key in enumerate(['custo_total', 'frete_total', 'frete_real_total', 'ganho_perda_liquido'], 3):
            cell = ws.cell(row=row, column=col)
            cell.value = round(float(c.get(key) or 0), 2)
            cell.number_format = FMT_MOEDA
            cell.border = thin_border
        row += 1

    # Embalagens por canal
    if canais:
        row += 2
        ws.cell(row=row, column=1, value='Embalagens utilizadas por canal').font = Font(bold=True, size=12)
        row += 1
        for c in canais:
            canal_nome = c.get('canal') or 'Não identificado'
            ws.cell(row=row, column=1, value=f'Canal: {canal_nome}').font = Font(bold=True)
            row += 1
            caixas = c.get('caixas') or []
            if caixas:
                for col, h in enumerate(['Embalagem', 'Quantidade', 'Custo unit. (R$)', 'Valor total (R$)'], 1):
                    cell = ws.cell(row=row, column=col, value=h)
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                    cell.fill = FILL_CABECALHO
                row += 1
                for cx in caixas:
                    ws.cell(row=row, column=1, value=cx.get('nome') or '-').border = thin_border
                    ws.cell(row=row, column=2, value=cx.get('quantidade') or 0).border = thin_border
                    c3 = ws.cell(row=row, column=3)
                    c3.value = round(float(cx.get('custo_unitario') or 0), 2)
                    c3.number_format = FMT_MOEDA
                    c3.border = thin_border
                    c4 = ws.cell(row=row, column=4)
                    c4.value = round(float(cx.get('valor_total') or 0), 2)
                    c4.number_format = FMT_MOEDA
                    c4.border = thin_border
                    row += 1
            else:
                ws.cell(row=row, column=1, value='Nenhuma embalagem registrada').border = thin_border
                row += 1
            row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.freeze_panes = 'A4'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def exportar_dashboard(data):
    """
    Gera planilha Excel reconstruindo os dados da Dashboard.
    data: dict retornado por GET /api/dashboard (pedidos_abertos, hoje, ontem, acumulado, por_canal, grafico_diario, custo_embalagem_total_geral, embalagens)
    Retorna: BytesIO com o arquivo .xlsx
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dashboard'

    hoje_str = (data.get('hoje') or {}).get('data', date.today().isoformat())
    ws['A1'] = f'Dashboard - Sistema de Logística ({hoje_str})'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    row = 3

    # Resumo (cards)
    ws.cell(row=row, column=1, value='Resumo da Dashboard').font = Font(bold=True, size=12)
    row += 1
    acumulado = data.get('acumulado') or {}
    media_diaria = acumulado.get('media_diaria') or 0
    pedidos_abertos = data.get('pedidos_abertos')
    pedidos_abertos = int(pedidos_abertos) if pedidos_abertos is not None else 0
    percentual_meta = round((media_diaria / pedidos_abertos * 100), 1) if pedidos_abertos > 0 else 0

    cards = [
        ('Acumulado Total', acumulado.get('total', 0)),
        ('Média Diária', media_diaria),
        ('Dia Anterior', (data.get('ontem') or {}).get('total_pedidos', 0)),
        ('Ideal Média Diária (meta)', pedidos_abertos),
        ('% Meta', percentual_meta),
        ('Pedidos em aberto', pedidos_abertos),
        ('Embalagens Usadas', (data.get('embalagens') or {}).get('usadas_mes', 0)),
        ('Custo Embalagem (Total)', data.get('custo_embalagem_total_geral')),
    ]
    for label, valor in cards:
        ws.cell(row=row, column=1, value=label).border = thin_border
        cell_b = ws.cell(row=row, column=2)
        if label == 'Custo Embalagem (Total)' and valor is not None:
            cell_b.value = round(float(valor), 2)
            cell_b.number_format = FMT_MOEDA
        elif label == '% Meta':
            cell_b.value = round(float(valor), 1)
            cell_b.number_format = '0.0"%"'
        else:
            cell_b.value = valor if valor is not None else 0
        cell_b.border = thin_border
        row += 1

    # Hoje
    row += 1
    h = data.get('hoje') or {}
    ws.cell(row=row, column=1, value='Hoje').font = Font(bold=True, size=12)
    row += 1
    for label, key in [
        ('Total pedidos', 'total_pedidos'),
        ('Custo total (R$)', 'custo_total'),
        ('Frete total / Pago cliente (R$)', 'frete_total'),
        ('Frete real (R$)', 'frete_real_total'),
        ('Ganho/Perda líquido (R$)', 'ganho_perda_liquido'),
    ]:
        ws.cell(row=row, column=1, value=label).border = thin_border
        v = h.get(key)
        cell_b = ws.cell(row=row, column=2)
        if key != 'total_pedidos' and v is not None:
            cell_b.value = round(float(v), 2)
            cell_b.number_format = FMT_MOEDA
        else:
            cell_b.value = v if v is not None else 0
        cell_b.border = thin_border
        row += 1

    # Por Canal
    por_canal = data.get('por_canal') or []
    if por_canal:
        row += 2
        ws.cell(row=row, column=1, value='Por Canal').font = Font(bold=True, size=12)
        row += 1
        headers_canal = ['Canal', 'Quantidade', 'Frete total (R$)', 'Frete real (R$)', 'Custo embalagem (R$)', 'Ganho/Perda (R$)']
        for col, h in enumerate(headers_canal, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = FILL_CABECALHO
        row += 1
        for c in por_canal:
            ws.cell(row=row, column=1, value=c.get('canal') or '-').border = thin_border
            ws.cell(row=row, column=2, value=c.get('quantidade') or 0).border = thin_border
            for col, key in enumerate(['frete_total', 'frete_real_total', 'custo_embalagem_total', 'ganho_perda_liquido'], 3):
                cell = ws.cell(row=row, column=col)
                cell.value = round(float(c.get(key) or 0), 2)
                cell.number_format = FMT_MOEDA
                cell.border = thin_border
            row += 1

    # Gráfico diário (pedidos por dia do mês)
    grafico = data.get('grafico_diario') or []
    if grafico:
        row += 2
        ws.cell(row=row, column=1, value='Pedidos por dia (mês)').font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value='Dia').font = Font(bold=True)
        ws.cell(row=row, column=2, value='Data').font = Font(bold=True)
        ws.cell(row=row, column=3, value='Quantidade').font = Font(bold=True)
        for c in [1, 2, 3]:
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = FILL_CABECALHO
        row += 1
        for d in grafico:
            ws.cell(row=row, column=1, value=d.get('dia')).border = thin_border
            ws.cell(row=row, column=2, value=d.get('data')).border = thin_border
            ws.cell(row=row, column=3, value=d.get('quantidade', 0)).border = thin_border
            row += 1

    # Embalagens utilizadas (detalhadas)
    emb = data.get('embalagens') or {}
    detalhadas = emb.get('detalhadas') or []
    row += 2
    ws.cell(row=row, column=1, value='Embalagens utilizadas').font = Font(bold=True, size=12)
    row += 1
    for col, h in enumerate(['Embalagem', 'Quantidade', 'Custo unit. (R$)', 'Valor total (R$)'], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = FILL_CABECALHO
    row += 1
    if detalhadas:
        for d in detalhadas:
            ws.cell(row=row, column=1, value=d.get('nome') or '-').border = thin_border
            ws.cell(row=row, column=2, value=d.get('quantidade') or 0).border = thin_border
            c3 = ws.cell(row=row, column=3)
            c3.value = round(float(d.get('custo_unitario') or 0), 2)
            c3.number_format = FMT_MOEDA
            c3.border = thin_border
            c4 = ws.cell(row=row, column=4)
            c4.value = round(float(d.get('valor_total') or 0), 2)
            c4.number_format = FMT_MOEDA
            c4.border = thin_border
            row += 1
        ws.cell(row=row, column=1, value='Total').font = Font(bold=True)
        ws.cell(row=row, column=2, value=emb.get('usadas_mes', 0)).font = Font(bold=True)
        ws.cell(row=row, column=4, value=round(float(emb.get('valor_total_mes') or 0), 2)).font = Font(bold=True)
        ws.cell(row=row, column=4).number_format = FMT_MOEDA
        for c in [1, 2, 3, 4]:
            ws.cell(row=row, column=c).border = thin_border
        row += 1
    else:
        ws.cell(row=row, column=1, value='Nenhuma embalagem utilizada no período.').border = thin_border
        row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.freeze_panes = 'A2'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
