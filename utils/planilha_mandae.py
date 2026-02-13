# -*- coding: utf-8 -*-
"""
Leitura de planilha Mandaê: colunas Código_Rastreio e Frete_Real.
Suporta CSV e Excel (.xlsx, .xls).
"""
import csv
import io
import re


def _normalizar_nome_coluna(nome):
    if not nome:
        return ''
    s = (nome or '').strip().lower()
    # Remove acentos comuns para comparação
    s = s.replace('ó', 'o').replace('ô', 'o').replace('ã', 'a').replace('á', 'a').replace('â', 'a')
    s = re.sub(r'[\s_]+', '_', s)
    return s


def _encontrar_indices_cabecalho(linha):
    """Dado uma lista de células do cabeçalho, retorna (idx_codigo_rastreio, idx_frete_real) ou (None, None)."""
    idx_codigo = None
    idx_frete = None
    for i, cell in enumerate(linha):
        val = (cell if isinstance(cell, str) else str(cell or '')).strip()
        norm = _normalizar_nome_coluna(val)
        if 'codigo' in norm and 'rastreio' in norm:
            idx_codigo = i
        if 'frete' in norm and 'real' in norm:
            idx_frete = i
    return idx_codigo, idx_frete


def ler_planilha_mandae(stream, filename=''):
    """
    Lê arquivo CSV ou Excel e retorna lista de dicts { 'codigo_rastreio': str, 'frete_real': float }.
    stream: file-like (bytes ou text para CSV).
    filename: usado para detectar extensão (.csv, .xlsx, .xls).
    """
    nome = (filename or '').lower()
    resultado = []

    if nome.endswith('.csv'):
        return _ler_csv_mandae(stream)
    if nome.endswith('.xlsx') or nome.endswith('.xls'):
        return _ler_excel_mandae(stream, nome)

    # Tenta como CSV por padrão
    return _ler_csv_mandae(stream)


def _ler_csv_mandae(stream):
    resultado = []
    content = stream.read()
    if isinstance(content, bytes):
        for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                text = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = content.decode('latin-1', errors='replace')
    else:
        text = content
    if not text.strip():
        return resultado
    first_line = text.split('\n')[0]
    sep = ';' if ';' in first_line else ','
    reader = csv.reader(io.StringIO(text), delimiter=sep)
    rows = list(reader)
    if not rows:
        return resultado
    header = rows[0]
    idx_codigo, idx_frete = _encontrar_indices_cabecalho(header)
    if idx_codigo is None or idx_frete is None:
        return resultado
    for row in rows[1:]:
        if len(row) <= max(idx_codigo, idx_frete):
            continue
        codigo = (row[idx_codigo] or '').strip()
        if not codigo:
            continue
        try:
            frete_val = float(str(row[idx_frete] or '0').replace(',', '.').strip())
        except (ValueError, TypeError):
            continue
        resultado.append({'codigo_rastreio': codigo, 'frete_real': round(frete_val, 2)})
    return resultado


def _ler_excel_mandae(stream, filename):
    resultado = []
    content = stream.read()
    if isinstance(content, str):
        content = content.encode('utf-8')
    # .xls (Excel 97-2003) com xlrd
    if filename.endswith('.xls'):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            sheet = wb.sheet_by_index(0)
            if sheet.nrows < 2:
                return resultado
            header = [sheet.cell_value(0, c) for c in range(sheet.ncols)]
            idx_codigo, idx_frete = _encontrar_indices_cabecalho(header)
            if idx_codigo is None or idx_frete is None:
                return resultado
            for r in range(1, sheet.nrows):
                codigo = (str(sheet.cell_value(r, idx_codigo)) or '').strip()
                if not codigo:
                    continue
                try:
                    frete_val = float(sheet.cell_value(r, idx_frete) or 0)
                except (ValueError, TypeError):
                    continue
                resultado.append({'codigo_rastreio': codigo, 'frete_real': round(frete_val, 2)})
            return resultado
        except ImportError:
            pass
        return resultado
    # .xlsx com openpyxl
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return resultado
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx_codigo, idx_frete = _encontrar_indices_cabecalho(header)
    if idx_codigo is None or idx_frete is None:
        wb.close()
        return resultado
    for row in ws.iter_rows(min_row=2):
        cells = [row[i].value if i < len(row) else None for i in range(max(idx_codigo, idx_frete) + 1)]
        if len(cells) <= max(idx_codigo, idx_frete):
            continue
        codigo = (str(cells[idx_codigo] or '') or '').strip()
        if not codigo:
            continue
        try:
            v = cells[idx_frete]
            frete_val = float(v) if v is not None else 0
        except (ValueError, TypeError):
            continue
        resultado.append({'codigo_rastreio': codigo, 'frete_real': round(frete_val, 2)})
    wb.close()
    return resultado
